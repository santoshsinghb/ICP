# utils/excel_reader.py
import openpyxl
from openpyxl import load_workbook


# ──────────────────────────────────────────
# Stand‑alone utilities
# ──────────────────────────────────────────
def read_excel_data(file_path, sheet_name):
    """
    Read an entire sheet into a list of dicts.
    Each dict = one row keyed by header row (row‑1) cells.
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    data = [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return data


def read_excel_with_filter(file_path, sheet_name, filter_by=None):
    """
    Same as read_excel_data but optionally filter by {column: value}.
    `filter_by` comparison is case‑insensitive and trimmed.
    """
    rows = read_excel_data(file_path, sheet_name)
    if not filter_by:
        return rows

    def _match(row):
        return all(
            str(row.get(k, "")).strip().lower() == str(v).strip().lower()
            for k, v in filter_by.items()
        )

    return [r for r in rows if _match(r)]


# ──────────────────────────────────────────
# Class wrapper (optional)
# ──────────────────────────────────────────
class excel:
    """Namespace wrapper so you can call Excel.get_login_data(...)."""

    @staticmethod
    def get_login_data(file_path, sheet_name):
        """Return ALL login rows."""
        return read_excel_data(file_path, sheet_name)

    @staticmethod
    def get_test_case(file_path, sheet_name, case_name):
        """Return **one** row where first column equals `case_name`."""
        for row in read_excel_data(file_path, sheet_name):
            if str(row[next(iter(row))]) == case_name:
                return row
        raise ValueError(f"Test‑case '{case_name}' not found in '{sheet_name}'")
